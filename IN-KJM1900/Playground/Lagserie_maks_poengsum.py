import numpy as np
from itertools import combinations
import openpyxl
import time
import operator as op
from functools import reduce

start = time.process_time()

def ncr(n, r):
    r = min(r, n-r)
    numer = reduce(op.mul, range(n, n-r, -1), 1)
    denom = reduce(op.mul, range(1, r+1), 1)
    return numer / denom


def Sort(x):
    l = len(x)
    for i in range(0, l):
        for j in range(0, l-i-1):
            if (int(x[j][0]) < int(x[j + 1][0])):
                tempo = x[j]
                x[j]= x[j + 1]
                x[j + 1]= tempo
    return x

def find_in_list_of_list(mylist, char):
    for sub_list in mylist:
        if char in sub_list:
            return (mylist.index(sub_list), sub_list.index(char))


#start = time.process_time()

WB = openpyxl.load_workbook('Lagserie-info.xlsx')

ark1 = WB.get_sheet_by_name('Ark1')

navn = []
øvelse = []
poeng = []

print(f"{time.process_time() - start:.1f}s")

for i in range(4,ark1.max_row+1):
    if isinstance(ark1.cell(row=i,column=16).value,str)==True:
        navn.append(ark1.cell(row=i,column=16).value)
        øvelse.append(ark1.cell(row=i,column=17).value)
        poeng.append(ark1.cell(row=i,column=18).value)

tekniske_øvelser = ['Høyde','Høyde u.t','Lengde u.t','Lengde','Tresteg','Stav','Kule','Slegge','Diskos','Spyd']
obligatoriske_øvelser = ['Høyde','Lengde','Tresteg','Stav','Kule','Slegge','Diskos','Spyd','100m','200m','400m','800m','1500m','3000mH','5000m','10000m','110mhk','400mhk']
løpsøvelser = ['100m','200m','400m','800m','1500m','3000m','3000mH','5000m','10000m','110mhk','400mhk']

ai = []; bi = []; ci = []

o = 0
for i in range(len(navn)):
    for u in obligatoriske_øvelser:
        if øvelse[i]==u:
            ai.append([])
            ai[o].append(poeng[i])
            ai[o].append(navn[i])
            ai[o].append(øvelse[i])
            o += 1
o = 0
for i in range(len(navn)):
    for u in tekniske_øvelser:
        if øvelse[i]==u:
            bi.append([])
            bi[o].append(int(poeng[i]))
            bi[o].append(navn[i])
            bi[o].append(øvelse[i])
            o += 1
o = 0
for i in range(len(navn)):
    for u in løpsøvelser:
        if øvelse[i]==u:
            ci.append([])
            ci[o].append(int(poeng[i]))
            ci[o].append(navn[i])
            ci[o].append(øvelse[i])
            o += 1



a = np.array(Sort(ai))

n_ = []
for i in a[:,2]:
    n_.append(find_in_list_of_list(a.tolist(),i)[0])
ok = []
for i in sorted(set(n_)):
    ok.append(a.tolist()[i])

while len(ok)<30:
    ok.append([0,'O','O'])
while len(Sort(bi))<30:
    bi.append([0,'O','O'])
while len(Sort(ci))<30:
    ci.append([0,'O','O'])

a = np.array(ok[:13]); b = np.array(Sort(bi)); c = np.array(Sort(ci))
AaA = np.array(ok[:13])


ok_alt = np.array(ai)
ok_uten_samme_øvelser = np.array(ai)

ok_ = []
ok_øvelser_ = []


for k1,k2,k3 in zip(AaA[:,1],AaA[:,2],AaA[:,0]):
    for m in range(len(ok_alt)):
        if k1==ok_alt[m,1] and k2==ok_alt[m,2] and k3==ok_alt[m,0]:
            ok_.append(m)
        if k2==ok_uten_samme_øvelser[m,2]:
            ok_øvelser_.append(m)

#print(ok_uten_samme_øvelser)
#print(sorted(ok_øvelser_))

print(sorted(ok_øvelser_))

ok_uten_samme_øvelser = ok_uten_samme_øvelser.tolist()
ok_alt = ok_alt.tolist()
for m in reversed(sorted(set(ok_))):
    del ok_alt[m]
for m in reversed(sorted(set(ok_øvelser_))):
    del ok_uten_samme_øvelser[m]
ok_ = []

if len(ok_uten_samme_øvelser)>0:
    ok_uten_samme_øvelser = np.array(ok_uten_samme_øvelser)
    for k1,k2,k3 in zip(ok_uten_samme_øvelser[:,1],ok_uten_samme_øvelser[:,2],ok_uten_samme_øvelser[:,0]):
        for m in range(len(ok_alt)):
            if k1==ok_alt[m][1] and k2==ok_alt[m][2] and k3==ok_alt[m][0]:
                ok_.append(m)
    ok_uten_samme_øvelser = ok_uten_samme_øvelser.tolist()
for m in reversed(ok_):
    del ok_alt[m]


#ok_uten_samme_øvelser -> trengs senere
#ok_alt -> trengs senere


tek_løp = 0
#-----
#print(ok_alt)
#print(ok_uten_samme_øvelser)
#print("--")
#print(AaA)


b = np.array(Sort(bi)); c = np.array(Sort(ci)); a = a[:13]

b_ = []; c_ = []
for k1,k2,k3 in zip(a[:,1],a[:,2],a[:,0]):
    for m in range(len(b)):
        if k1==b[m,1] and k2==b[m,2] and k3==b[m,0]:
            b_.append(m)
    for m in range(len(c)):
        if k1==c[m,1] and k2==c[m,2] and k3==c[m,0]:
            c_.append(m)


b = b.tolist();    c = c.tolist(); a = a.tolist()
for m in reversed(sorted(set(b_))):
    del b[m]
for m in reversed(sorted(set(c_))):
    del c[m]

b_and_c = np.array(Sort(b + c))

count_i = 0
for i in b_and_c[:13+tek_løp,2]:
    if not 'm' in i:
        count_i += 1
b_and_c = b_and_c.tolist()

#print(all_arr)
#print("Ok uten samme øvelser")
#print(ok_uten_samme_øvelser)

if count_i >= 4+tek_løp:
    all_arr = np.array(a+b_and_c[:12])
else:
    all_arr = np.array((a+b[:4+tek_løp]+c[:8]))


overlapp_øvelser = []
for i in set(all_arr[:,1]):

    b = np.array(Sort(bi)); c = np.array(Sort(ci)); a = np.array(a)

    b_ = []; c_ = []
    for k1,k2,k3 in zip(a[:,1],a[:,2],a[:,0]):
        for m in range(len(b)):
            if k1==b[m,1] and k2==b[m,2] and k3==b[m,0]:
                b_.append(m)
        for m in range(len(c)):
            if k1==c[m,1] and k2==c[m,2] and k3==c[m,0]:
                c_.append(m)

    b = b.tolist();    c = c.tolist(); a = a.tolist()
    for m in reversed(sorted(set(b_))):
        del b[m]
    for m in reversed(sorted(set(c_))):
        del c[m]
    overlapp_øvelser.append(i)
    overlapp_øvelser.append(sum(x.count(i) for x in all_arr.tolist()))
    if sum(x.count(i) for x in all_arr.tolist())>5 and i!='O':
        aba = []
        oi = sum(x.count(i) for x in all_arr.tolist())
        for new_aba in all_arr.tolist():
            if new_aba[1]==i:
                aba.append(new_aba)
        for y in range(1,sum(x.count(i) for x in all_arr.tolist())-4):
            #print(aba[4+y])
            print("!")
            tek_løp += 1
            if isinstance(find_in_list_of_list(ok_alt,aba[-y-1][2]),tuple)==True:
                a.append(ok_alt[(find_in_list_of_list(ok_alt,aba[-y-1][2])[0])])
                print(ok_alt[(find_in_list_of_list(ok_alt,aba[-y-1][2])[0])])
                del ok_alt[(find_in_list_of_list(ok_alt,aba[-y-1][2])[0])]
            if len(ok_uten_samme_øvelser)>0:
                a.append(ok_uten_samme_øvelser[0])
                all_arr.tolist().append(ok_uten_samme_øvelser[0])
                print(ok_uten_samme_øvelser[0])
                del ok_uten_samme_øvelser[0]
        for m in range(5):
            if isinstance(find_in_list_of_list(ok_alt,aba[m][2]),tuple)==True:
                a.append(ok_alt[(find_in_list_of_list(ok_alt,aba[m][2])[0])])
                print(ok_alt[(find_in_list_of_list(ok_alt,aba[m][2])[0])])
                del ok_alt[(find_in_list_of_list(ok_alt,aba[m][2])[0])]


for i in range(20):
    if count_i >= 4+tek_løp:
        all_arr = np.array(a+b_and_c[:12+tek_løp])
    else:
        all_arr = np.array((a+b[:4+tek_løp]+c[:8+tek_løp]))

    for j,k in zip(range(1,len(overlapp_øvelser),2),set(all_arr[:,1])):
        #print(sum(x.count(k) for x in all_arr.tolist()))
        #print(overlapp_øvelser[j])
        if sum(x.count(k) for x in all_arr.tolist())>5 and overlapp_øvelser[j]<sum(x.count(k) for x in all_arr.tolist()) and i!='O':
            print(k,sum(x.count(k) for x in all_arr.tolist()))

            aba = []
            for new_aba in all_arr.tolist():
                if new_aba[1]==k:
                    aba.append(new_aba)
            #print(aba[4+y])
            #print(aba)
            #print(len(aba))

            for m in range(len(all_arr.tolist())):
                if all_arr.tolist()[m]==aba[-i-1]:
                    tek_løp += 1
            if isinstance(find_in_list_of_list(ok_alt,aba[i-1][2]),tuple)==True:
                a.append(ok_alt[(find_in_list_of_list(ok_alt,aba[-i-1][2])[0])])
                del ok_alt[(find_in_list_of_list(ok_alt,aba[-i-1][2])[0])]
            if len(ok_uten_samme_øvelser)>0:
                a.append(ok_uten_samme_øvelser[0])
                del ok_uten_samme_øvelser[0]
    overlapp_øvelser = []
    overlapp_øvelser.append(i)
    overlapp_øvelser.append(sum(x.count(i) for x in all_arr.tolist()))

    h = 1
    for j,k in zip(range(1,len(overlapp_øvelser),2),set(all_arr[:,1])):
        if sum(x.count(k) for x in all_arr.tolist())>5 and overlapp_øvelser[j]<sum(x.count(k) for x in all_arr.tolist()):
            h = 0
            break
    if h==1:
        break
    else:
        continue

a = np.array(Sort(a)); b = np.array(Sort(bi)); c = np.array(Sort(ci))
#print(c)

# må legge til ekstre dersom nye lagt til navn har mer enn 5 og det de tidligere hadde

counter2 = 0
counter3 = 0
counter = 0
n = 0

#print(a)
#print(b)
#print(c)

#print(tek_løp)
print(len(a))

combos = ncr(len(a),13)


#print(a)
#print(b)
#print(c)




for combination1,øvelse1,navn1 in zip(combinations(a[:,0].astype(int), 13),combinations(a[:,2], 13),combinations(a[:,1], 13)):

    b = np.array(Sort(bi));   c = np.array(Sort(ci))
    counter2 += 1
    if counter2 % 10000==0:
        print(f"Unnagjort: {counter2/10000:.0f} av {combos/10000:.0f} - Estimert tid igjen: {(combos/counter2*((time.process_time() - start))-(time.process_time() - start))/250:.1f}min")
    #----
    oop = 1
    for i in set(navn1):
        if (sum(x.count(i) for x in navn1))>5:
            oop = 0
            break
    if oop==0:
        continue
    #---

    for m in range(18):
        o = 1
        index = øvelse1.count(obligatoriske_øvelser[m])
        if index >=2:
            o = 0
            #print(m)
            #print(øvelse1)
            break
    if o ==0:
        continue

    b_ = []; c_ = []
    for k1,k2 in zip(øvelse1,combination1):
        for m in range(len(b)):
            if k1==b[m,2] and k2<=b[m,0].astype(int):
                b_.append(m)
        for m in range(len(c)):
            if k1==c[m,2] and k2<=c[m,0].astype(int):
                c_.append(m)

    #for k1,k2,k3 in zip(navn1,øvelse1,combination1):
     #   for m in range(len(b)):
      #      if k1==b[m,1] and k2==b[m,2] and k3==b[m,0].astype(int):
       #         b_.append(m)
        #for m in range(len(c)):
         #   if k1==c[m,1] and k2==c[m,2] and k3==c[m,0].astype(int):
          #      c_.append(m)

    b = b.tolist();    c = c.tolist()
    for m in reversed(sorted(set(b_))):
        del b[m]
    for m in reversed(sorted(set(c_))):
        del c[m]
    #print(combination1)

    b_and_c = np.array(b_and_c)
    if sum(b_and_c[:12,0].astype(int))+sum(combination1)<=n:
        continue

    count_i = 0
    for i in b_and_c[:13+tek_løp,2]:
        if not 'm' in i:
            count_i += 1
    b_and_c = b_and_c.tolist()

    counter3 += 1
    print(f"{counter3} - {counter2}/{combos:.0f}")#*(counter2/combos)):.0f

    okeey = 0
    if count_i >= 4+tek_løp:
        b = np.array(b[:(count_i+tek_løp)]); c = np.array(c[:(12-count_i+tek_løp)])
        for combination2,øvelse2,navn2 in zip(combinations(b[:,0].astype(int), count_i),combinations(b[:,2], count_i),combinations(b[:,1], count_i)):
            for combination3,øvelse3,navn3 in zip(combinations(c[:,0].astype(int), 12-count_i),combinations(c[:,2], 12-count_i),combinations(c[:,1], 12-count_i)):
                counter += 1
                oop = 1
                for i in set(navn1+navn2+navn3):
                    if (sum(x.count(i) for x in navn1+navn2+navn3))>5:
                        oop = 0
                        break
                if oop==0:
                    continue
                if sum(combination1+combination2+combination3)>okeey:
                    okeey = sum(combination1+combination2+combination3)
                if sum(combination1+combination2+combination3)>n:
                     n = sum(combination1+combination2+combination3)
                     print((f"Ny funnet makspoengsum: {sum(combination1+combination2+combination3)}"))
                     comb1 = combination1; comb2 = combination2; comb3 = combination3
                     øv1 = øvelse1; øv2 = øvelse2; øv3 = øvelse3
                     na1 = navn1; na2 = navn2; na3 = navn3

    else:
        b = np.array(b[:(4+tek_løp)]); c = np.array(c[:(8+tek_løp)])
        for combination2,øvelse2,navn2 in zip(combinations(b[:,0].astype(int), 4),combinations(b[:,2], 4),combinations(b[:,1], 4)):
            for combination3,øvelse3,navn3 in zip(combinations(c[:,0].astype(int), 8),combinations(c[:,2], 8),combinations(c[:,1], 8)):
                counter += 1
                oop = 1
                for i in set(navn1+navn2+navn3):
                    if (sum(x.count(i) for x in navn1+navn2+navn3))>5:
                        oop = 0
                        break
                if oop==0:
                    continue
                if sum(combination1+combination2+combination3)>okeey:
                    okeey = sum(combination1+combination2+combination3)
                if sum(combination1+combination2+combination3)>n:
                     n = sum(combination1+combination2+combination3)
                     print((f"Ny funnet makspoengsum: {sum(combination1+combination2+combination3)}"))
                     comb1 = combination1; comb2 = combination2; comb3 = combination3
                     øv1 = øvelse1; øv2 = øvelse2; øv3 = øvelse3
                     na1 = navn1; na2 = navn2; na3 = navn3
    print(okeey)

print(f"{time.process_time() - start:.1f}s")


print("")
print("")
print("------")
print(f"Kombinasjoner sjekket: {counter}")
print("")
print(f"{time.process_time() - start:.1f}s")
print("")
print("Obligatoriske Øvelser")
print("---------------------")
for i in range(len(comb1)):
    print(comb1[i],na1[i],øv1[i])
print("")
print("Valgfri Tekniske Øvelser")
print("------------------------")
for i in range(len(comb2)):
    print(comb2[i],na2[i],øv2[i])
print("")
print("Valgfri Løpsøvelser")
print("-------------------")
for i in range(len(comb3)):
    print(comb3[i],na3[i],øv3[i])
print("")
print(f"Maksimal Poengsum: {n}")
